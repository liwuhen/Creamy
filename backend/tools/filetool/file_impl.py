import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def expansion_write_excel(data: list, output_path: str) -> str:
    """Generate an Excel table with formatting"""
    try:
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "零件库存盘点"

        # 表头
        headers = ["零件名称", "规格", "品牌", "材质", "当前库存", "状态", "最后更新时间"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        fill_normal = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        fill_warning = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        fill_critical = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row_idx, item in enumerate(data, 2):
            total = item.get("total", 0)
            status = "缺货" if total == 0 else "正常"

            ws.cell(row=row_idx, column=1, value=item.get("name", ""))
            ws.cell(row=row_idx, column=2, value=item.get("spec", ""))
            ws.cell(row=row_idx, column=3, value=item.get("brand", ""))
            ws.cell(row=row_idx, column=4, value=item.get("material", ""))
            ws.cell(row=row_idx, column=5, value=total)
            ws.cell(row=row_idx, column=6, value=status)
            ws.cell(row=row_idx, column=7, value=updated_at)

            if "缺货" in status:
                row_fill = fill_critical
            elif "不足" in status:
                row_fill = fill_warning
            else:
                row_fill = fill_normal

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = row_fill

        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(output_path)
        return f"Excel 已生成：{output_path}，共 {len(data)} 条记录"

    except Exception as e:
        return f"Excel 生成失败：{e!s}"
